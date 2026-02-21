
from openpyxl import load_workbook
from shared.mcp.protocol import ToolResult, TextContent
from shared.logging.main import get_logger
import os
import json
import pandas as pd

logger = get_logger(__name__)

def dict_to_result(data: dict) -> ToolResult:
    return ToolResult(content=[TextContent(text=json.dumps(data, indent=2))])

def _get_wb(path):
    if not os.path.exists(path):
        raise FileNotFoundError(f"File not found: {path}")
    return load_workbook(path)

async def clean_whitespace(arguments: dict) -> ToolResult:
    """Trim leading/trailing whitespace from strings in range."""
    try:
        path = arguments['file_path']
        sheet_name = arguments.get('sheet_name')
        range_string = arguments.get('range_string') # Optional, default used range
        
        wb = _get_wb(path)
        ws = wb[sheet_name] if sheet_name else wb.active
        
        count = 0
        rows = ws[range_string] if range_string else ws.iter_rows()
        
        for row in rows:
            for cell in row:
                if isinstance(cell.value, str):
                    original = cell.value
                    cleaned = original.strip()
                    if original != cleaned:
                        cell.value = cleaned
                        count += 1
                        
        wb.save(path)
        wb.close()
        return dict_to_result({"status": "whitespace_cleaned", "cells_modified": count})
    except Exception as e:
        return ToolResult(isError=True, content=[TextContent(text=str(e))])

async def remove_duplicates(arguments: dict) -> ToolResult:
    """
    Remove duplicate rows based on subset of columns.
    Uses Pandas for logic, then rewrites sheet data.
    """
    try:
        path = arguments['file_path']
        sheet_name = arguments.get('sheet_name')
        subset = arguments.get('subset') # List of column names or indices
        header_row = arguments.get('header_row', 1)
        
        # Load via Pandas for easy dedupe
        df = pd.read_excel(path, sheet_name=sheet_name, header=header_row-1)
        original_len = len(df)
        
        df.drop_duplicates(subset=subset, inplace=True)
        new_len = len(df)
        
        # Write back? This wipes formatting unless we be careful.
        # "Smart" rewrite: clear data range, write values back.
        # Just writing values is safer for preserving cell styles if we iterate.
        # But efficiently:
        
        if original_len == new_len:
             return dict_to_result({"status": "no_duplicates_found"})
             
        wb = _get_wb(path)
        ws = wb[sheet_name] if sheet_name else wb.active
        
        # Clear existing data part (keep headers)
        start_row = header_row + 1
        # Better to delete rows and insert?
        # Safe approach: Clear all value cells below header, write new data.
        
        # 1. Clear old data
        # Only clear up to max row of old data
        for row in ws.iter_rows(min_row=start_row):
            for cell in row:
                cell.value = None
                
        # 2. Write new data
        # Utilize existing write_tool logic or direct
        from openpyxl.utils.dataframe import dataframe_to_rows
        # This includes index and header usually, check args
        rows = dataframe_to_rows(df, index=False, header=False)
        
        r_idx = start_row
        for r_data in rows:
            for c_idx, val in enumerate(r_data, 1):
                ws.cell(row=r_idx, column=c_idx, value=val)
            r_idx += 1
            
        wb.save(path)
        wb.close()
        return dict_to_result({"status": "duplicates_removed", "removed_count": original_len - new_len})
    except Exception as e:
        return ToolResult(isError=True, content=[TextContent(text=str(e))])

async def convert_to_numbers(arguments: dict) -> ToolResult:
    """Attempt to convert string-numbers to floats/ints."""
    try:
        path = arguments['file_path']
        sheet_name = arguments.get('sheet_name')
        
        wb = _get_wb(path)
        ws = wb[sheet_name] if sheet_name else wb.active
        
        count = 0
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str):
                    try:
                        # Try int first, then float
                        if cell.value.isdigit():
                            cell.value = int(cell.value)
                            count += 1
                        else:
                            val = float(cell.value)
                            cell.value = val
                            count += 1
                    except:
                        pass
        
        wb.save(path)
        wb.close()
        return dict_to_result({"status": "converted_to_numbers", "count": count})
    except Exception as e:
        return ToolResult(isError=True, content=[TextContent(text=str(e))])
