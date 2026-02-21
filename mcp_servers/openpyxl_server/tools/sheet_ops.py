
from openpyxl import load_workbook
from shared.mcp.protocol import ToolResult, TextContent
from shared.logging.main import get_logger
import os
import json

logger = get_logger(__name__)

def dict_to_result(data: dict) -> ToolResult:
    return ToolResult(content=[TextContent(text=json.dumps(data, indent=2))])

def _get_wb(path, read_only=False, data_only=False):
    if not os.path.exists(path):
        raise FileNotFoundError(f"File not found: {path}")
    return load_workbook(path, read_only=read_only, data_only=data_only)

# ============================================================================
# Sheet Operations
# ============================================================================

async def list_worksheets(arguments: dict) -> ToolResult:
    try:
        path = arguments['file_path']
        wb = _get_wb(path, read_only=True)
        sheets = wb.sheetnames
        wb.close()
        return dict_to_result({"sheets": sheets, "count": len(sheets)})
    except Exception as e:
        return ToolResult(isError=True, content=[TextContent(text=str(e))])

async def create_worksheet(arguments: dict) -> ToolResult:
    try:
        path = arguments['file_path']
        title = arguments['title']
        index = arguments.get('index') # Optional int
        
        wb = _get_wb(path)
        ws = wb.create_sheet(title=title, index=index)
        wb.save(path)
        wb.close()
        return dict_to_result({"status": "created", "sheet": ws.title})
    except Exception as e:
        return ToolResult(isError=True, content=[TextContent(text=str(e))])

async def delete_worksheet(arguments: dict) -> ToolResult:
    try:
        path = arguments['file_path']
        sheet_name = arguments['sheet_name']
        
        wb = _get_wb(path)
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]
            wb.save(path)
            wb.close()
            return dict_to_result({"status": "deleted", "sheet": sheet_name})
        else:
            wb.close()
            return ToolResult(isError=True, content=[TextContent(text=f"Sheet '{sheet_name}' not found")])
    except Exception as e:
        return ToolResult(isError=True, content=[TextContent(text=str(e))])

async def rename_worksheet(arguments: dict) -> ToolResult:
    try:
        path = arguments['file_path']
        old_name = arguments['old_name']
        new_name = arguments['new_name']
        
        wb = _get_wb(path)
        if old_name in wb.sheetnames:
            ws = wb[old_name]
            ws.title = new_name
            wb.save(path)
            wb.close()
            return dict_to_result({"status": "renamed", "old": old_name, "new": new_name})
        else:
            wb.close()
            return ToolResult(isError=True, content=[TextContent(text=f"Sheet '{old_name}' not found")])
    except Exception as e:
        return ToolResult(isError=True, content=[TextContent(text=str(e))])

async def copy_worksheet(arguments: dict) -> ToolResult:
    try:
        path = arguments['file_path']
        source_name = arguments['source_sheet']
        target_title = arguments.get('new_title')
        
        wb = _get_wb(path)
        if source_name in wb.sheetnames:
            source = wb[source_name]
            target = wb.copy_worksheet(source)
            if target_title:
                target.title = target_title
            
            wb.save(path)
            final_name = target.title
            wb.close()
            return dict_to_result({"status": "copied", "from": source_name, "to": final_name})
        else:
            wb.close()
            return ToolResult(isError=True, content=[TextContent(text=f"Sheet '{source_name}' not found")])
    except Exception as e:
        return ToolResult(isError=True, content=[TextContent(text=str(e))])

async def get_sheet_properties(arguments: dict) -> ToolResult:
    try:
        path = arguments['file_path']
        sheet_name = arguments['sheet_name']
        
        wb = _get_wb(path) # Need full load for properties sometimes
        if sheet_name not in wb.sheetnames:
             return ToolResult(isError=True, content=[TextContent(text=f"Sheet not found")])
             
        ws = wb[sheet_name]
        props = {
            "title": ws.title,
            "sheet_state": ws.sheet_state, # visible, hidden, veryHidden
            "tabColor": ws.sheet_properties.tabColor.rgb if ws.sheet_properties.tabColor else None,
            "page_setup": {
                "orientation": ws.page_setup.orientation,
                "paperSize": ws.page_setup.paperSize,
            }
        }
        wb.close()
        return dict_to_result(props)
    except Exception as e:
        return ToolResult(isError=True, content=[TextContent(text=str(e))])
        
async def set_sheet_properties(arguments: dict) -> ToolResult:
    try:
        path = arguments['file_path']
        sheet_name = arguments['sheet_name']
        tab_color = arguments.get('tab_color') # RGB Hex
        state = arguments.get('state') # "visible", "hidden"
        
        wb = _get_wb(path)
        ws = wb[sheet_name]
        
        if tab_color:
            ws.sheet_properties.tabColor = tab_color
        
        if state:
            ws.sheet_state = state
            
        wb.save(path)
        wb.close()
        return dict_to_result({"status": "updated", "sheet": sheet_name})
    except Exception as e:
        return ToolResult(isError=True, content=[TextContent(text=str(e))])
