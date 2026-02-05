
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, Protection
from shared.mcp.protocol import ToolResult, TextContent
from shared.logging import get_logger
import os
import json

logger = get_logger(__name__)

def dict_to_result(data: dict) -> ToolResult:
    return ToolResult(content=[TextContent(text=json.dumps(data, indent=2))])

def _get_wb(path):
    if not os.path.exists(path):
        raise FileNotFoundError(f"File not found: {path}")
    return load_workbook(path)

async def set_cell_font(arguments: dict) -> ToolResult:
    """Set Font properties."""
    try:
        path = arguments['file_path']
        coord = arguments['cell']
        name = arguments.get('font_name')
        size = arguments.get('size')
        bold = arguments.get('bold')
        italic = arguments.get('italic')
        color = arguments.get('color') # RGB hex
        sheet_name = arguments.get('sheet_name')
        
        wb = _get_wb(path)
        ws = wb[sheet_name] if sheet_name else wb.active
        
        cell = ws[coord]
        # Current font props or default
        current = cell.font
        
        new_font = Font(
            name=name if name is not None else current.name,
            size=size if size is not None else current.sz,
            bold=bold if bold is not None else current.b,
            italic=italic if italic is not None else current.i,
            color=color if color is not None else current.color
        )
        cell.font = new_font
        
        wb.save(path)
        wb.close()
        return dict_to_result({"status": "font_updated", "cell": coord})
    except Exception as e:
        return ToolResult(isError=True, content=[TextContent(text=str(e))])

async def set_cell_fill(arguments: dict) -> ToolResult:
    """Set Background Fill (Solid)."""
    try:
        path = arguments['file_path']
        coord = arguments['cell']
        color = arguments['color'] # RGB Hex, e.g. 'FF0000'
        sheet_name = arguments.get('sheet_name')
        
        wb = _get_wb(path)
        ws = wb[sheet_name] if sheet_name else wb.active
        
        fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
        ws[coord].fill = fill
        
        wb.save(path)
        wb.close()
        return dict_to_result({"status": "fill_updated", "cell": coord, "color": color})
    except Exception as e:
        return ToolResult(isError=True, content=[TextContent(text=str(e))])

async def set_cell_border(arguments: dict) -> ToolResult:
    """Set Border (Thin/Thick/Double etc.)"""
    try:
        path = arguments['file_path']
        coord = arguments['cell']
        style = arguments.get('style', 'thin')
        color = arguments.get('color', '000000')
        sides = arguments.get('sides', ['left', 'right', 'top', 'bottom']) # List of sides
        sheet_name = arguments.get('sheet_name')
        
        wb = _get_wb(path)
        ws = wb[sheet_name] if sheet_name else wb.active
        
        side_obj = Side(border_style=style, color=color)
        border_args = {}
        if 'left' in sides: border_args['left'] = side_obj
        if 'right' in sides: border_args['right'] = side_obj
        if 'top' in sides: border_args['top'] = side_obj
        if 'bottom' in sides: border_args['bottom'] = side_obj
        
        new_border = Border(**border_args)
        ws[coord].border = new_border
        
        wb.save(path)
        wb.close()
        return dict_to_result({"status": "border_updated", "cell": coord})
    except Exception as e:
        return ToolResult(isError=True, content=[TextContent(text=str(e))])

async def merge_cells(arguments: dict) -> ToolResult:
    try:
        path = arguments['file_path']
        range_string = arguments['range_string']
        sheet_name = arguments.get('sheet_name')
        
        wb = _get_wb(path)
        ws = wb[sheet_name] if sheet_name else wb.active
        ws.merge_cells(range_string)
        wb.save(path)
        wb.close()
        return dict_to_result({"status": "merged", "range": range_string})
    except Exception as e:
        return ToolResult(isError=True, content=[TextContent(text=str(e))])

async def unmerge_cells(arguments: dict) -> ToolResult:
    try:
        path = arguments['file_path']
        range_string = arguments['range_string']
        sheet_name = arguments.get('sheet_name')
        
        wb = _get_wb(path)
        ws = wb[sheet_name] if sheet_name else wb.active
        ws.unmerge_cells(range_string)
        wb.save(path)
        wb.close()
        return dict_to_result({"status": "unmerged", "range": range_string})
    except Exception as e:
        return ToolResult(isError=True, content=[TextContent(text=str(e))])
