
from openpyxl import load_workbook, Workbook
from shared.mcp.protocol import ToolResult, TextContent
from shared.logging import get_logger
import os
import json

logger = get_logger(__name__)

def dict_to_result(data: dict, title: str = "Result") -> ToolResult:
    return ToolResult(content=[TextContent(text=json.dumps(data, indent=2, default=str))])

# ============================================================================
# 1. Multitalent / Super Tools
# ============================================================================

async def analyze_workbook_file(arguments: dict) -> ToolResult:
    """
    Multitalent: Deep scan of a workbook structure.
    Returns sheets, dimensions, defined names, charts count, and properties.
    """
    try:
        file_path = arguments['file_path']
        if not os.path.exists(file_path):
             return ToolResult(isError=True, content=[TextContent(text=f"File not found: {file_path}")])

        # Read only for speed
        wb = load_workbook(file_path, read_only=True, data_only=True)
        
        info = {
            "file_path": file_path,
            "sheet_names": wb.sheetnames,
            "named_ranges": list(wb.defined_names.keys()),
            "properties": {
                "author": wb.properties.creator,
                "created": wb.properties.created,
                "modified": wb.properties.modified,
            },
            "sheets": []
        }
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            sheet_info = {
                "name": sheet_name,
                "max_row": ws.max_row, # Can be inaccurate in read_only
                "max_column": ws.max_column,
            }
            # Approximate dimensions if max_row is None (common in read-only)
            if sheet_info['max_row'] is None:
                 sheet_info['max_row'] = "Unknown (Read-only mode)"
            
            info['sheets'].append(sheet_info)
            
        wb.close()
        return dict_to_result(info, "Workbook Analysis")

    except Exception as e:
        logger.error(f"Analysis failed: {e}")
        return ToolResult(isError=True, content=[TextContent(text=str(e))])


# ============================================================================
# 2. Workbook Lifecycle
# ============================================================================

async def create_new_workbook(arguments: dict) -> ToolResult:
    """Create a new empty .xlsx file."""
    try:
        file_path = arguments['file_path']
        if os.path.exists(file_path):
            overwrite = arguments.get('overwrite', False)
            if not overwrite:
                 return ToolResult(isError=True, content=[TextContent(text=f"File exists: {file_path}. Set overwrite=True to replace.")])
        
        wb = Workbook()
        wb.save(file_path)
        wb.close()
        return dict_to_result({"status": "created", "path": file_path})
        
    except Exception as e:
        return ToolResult(isError=True, content=[TextContent(text=str(e))])

async def get_workbook_metadata(arguments: dict) -> ToolResult:
    """Get core properties (author, dates)."""
    try:
        file_path = arguments['file_path']
        wb = load_workbook(file_path, read_only=True)
        props = {
            "title": wb.properties.title,
            "subject": wb.properties.subject,
            "author": wb.properties.creator,
            "created": wb.properties.created,
            "modified": wb.properties.modified,
            "lastModifiedBy": wb.properties.lastModifiedBy,
        }
        wb.close()
        return dict_to_result(props)
    except Exception as e:
        return ToolResult(isError=True, content=[TextContent(text=str(e))])

async def list_named_ranges(arguments: dict) -> ToolResult:
    """List all defined names (Named Ranges) and their targets."""
    try:
        file_path = arguments['file_path']
        wb = load_workbook(file_path, read_only=True)
        
        names = []
        for name, obj in wb.defined_names.items():
            # obj.value often contains the range like 'Sheet1!$A$1:$C$5'
            # Or obj.destinations generator
            targets = []
            if hasattr(obj, 'destinations'):
                for title, coord in obj.destinations:
                    targets.append(f"{title}!{coord}")
            
            names.append({
                "name": name,
                "value": obj.value if hasattr(obj, 'value') else None,
                "targets": targets
            })
            
        wb.close()
        return dict_to_result({"named_ranges": names})
    except Exception as e:
        return ToolResult(isError=True, content=[TextContent(text=str(e))])
