
from openpyxl import load_workbook
from shared.mcp.protocol import ToolResult, TextContent
from shared.logging.main import get_logger
import os
import json

logger = get_logger(__name__)

def dict_to_result(data: dict) -> ToolResult:
    return ToolResult(content=[TextContent(text=json.dumps(data, indent=2))])

def _get_wb(path, data_only=False):
    if not os.path.exists(path):
        raise FileNotFoundError(f"File not found: {path}")
    return load_workbook(path, data_only=data_only)

async def find_error_cells(arguments: dict) -> ToolResult:
    """Find cells with error values (#N/A, #REF!, etc)."""
    try:
        path = arguments['file_path']
        sheet_name = arguments.get('sheet_name')
        
        wb = _get_wb(path, data_only=True)
        ws = wb[sheet_name] if sheet_name else wb.active
        
        errors = []
        error_codes = ["#N/A", "#REF!", "#NAME?", "#DIV/0!", "#NULL!", "#VALUE!", "#NUM!"]
        
        for row in ws.iter_rows():
            for cell in row:
                if cell.value in error_codes:
                    errors.append({
                        "cell": cell.coordinate,
                        "error": cell.value
                    })
                    
        wb.close()
        return dict_to_result({"errors": errors, "count": len(errors)})
    except Exception as e:
        return ToolResult(isError=True, content=[TextContent(text=str(e))])

async def list_all_formulas(arguments: dict) -> ToolResult:
    try:
        path = arguments['file_path']
        sheet_name = arguments.get('sheet_name')
        
        wb = _get_wb(path, data_only=False)
        ws = wb[sheet_name] if sheet_name else wb.active
        
        formulas = []
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    formulas.append({
                        "cell": cell.coordinate,
                        "formula": cell.value
                    })
        wb.close()
        return dict_to_result({"formulas": formulas, "count": len(formulas)})
    except Exception as e:
        return ToolResult(isError=True, content=[TextContent(text=str(e))])
