
from openpyxl import load_workbook
from shared.mcp.protocol import ToolResult, TextContent
from shared.logging.main import get_logger
import os
import json

logger = get_logger(__name__)

def dict_to_result(data: dict) -> ToolResult:
    return ToolResult(content=[TextContent(text=json.dumps(data, indent=2))])

def _get_wb(path):
    if not os.path.exists(path):
        raise FileNotFoundError(f"File not found: {path}")
    return load_workbook(path) # Read/Write mode

# ============================================================================
# Write Operations
# ============================================================================

async def write_cell_value(arguments: dict) -> ToolResult:
    try:
        path = arguments['file_path']
        sheet_name = arguments.get('sheet_name')
        coord = arguments['cell']
        value = arguments['value']
        
        wb = _get_wb(path)
        ws = wb[sheet_name] if sheet_name else wb.active
        
        ws[coord] = value
        
        wb.save(path)
        final_val = ws[coord].value
        wb.close()
        return dict_to_result({"status": "written", "cell": coord, "value": final_val})
    except Exception as e:
        return ToolResult(isError=True, content=[TextContent(text=str(e))])

async def write_range_values(arguments: dict) -> ToolResult:
    """
    Bulk write.
    Args:
        data: List of lists [[1,2], [3,4]]
        start_cell: 'A1'
    """
    try:
        path = arguments['file_path']
        sheet_name = arguments.get('sheet_name')
        start_cell = arguments.get('start_cell', 'A1')
        data = arguments['data']
        
        wb = _get_wb(path)
        ws = wb[sheet_name] if sheet_name else wb.active
        
        # Calculate start row/col
        from openpyxl.utils.cell import coordinate_to_tuple
        start_row, start_col = coordinate_to_tuple(start_cell)
        
        for r_idx, row_data in enumerate(data):
            for c_idx, value in enumerate(row_data):
                ws.cell(row=start_row + r_idx, column=start_col + c_idx, value=value)
                
        wb.save(path)
        wb.close()
        return dict_to_result({"status": "bulk_written", "rows": len(data), "cols": len(data[0]) if data else 0})
    except Exception as e:
        return ToolResult(isError=True, content=[TextContent(text=str(e))])

async def write_formula(arguments: dict) -> ToolResult:
    try:
        path = arguments['file_path']
        cell = arguments['cell']
        formula = arguments['formula'] # "=SUM(A1:A5)"
        sheet_name = arguments.get('sheet_name')
        
        wb = _get_wb(path)
        ws = wb[sheet_name] if sheet_name else wb.active
        
        ws[cell] = formula
        wb.save(path)
        wb.close()
        return dict_to_result({"status": "formula_written", "cell": cell, "formula": formula})
    except Exception as e:
        return ToolResult(isError=True, content=[TextContent(text=str(e))])

async def clear_range(arguments: dict) -> ToolResult:
    try:
        path = arguments['file_path']
        range_string = arguments['range_string']
        sheet_name = arguments.get('sheet_name')
        
        wb = _get_wb(path)
        ws = wb[sheet_name] if sheet_name else wb.active
        
        # ws[range] returns tuple of tulples
        for row in ws[range_string]:
            for cell in row:
                cell.value = None
                # Clear style too? Probably safer not to, just value
                
        wb.save(path)
        wb.close()
        return dict_to_result({"status": "cleared", "range": range_string})
    except Exception as e:
        return ToolResult(isError=True, content=[TextContent(text=str(e))])
