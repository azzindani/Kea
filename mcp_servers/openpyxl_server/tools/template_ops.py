
from openpyxl import load_workbook
from shared.mcp.protocol import ToolResult, TextContent
from shared.logging import get_logger
import os
import json

logger = get_logger(__name__)

def dict_to_result(data: dict) -> ToolResult:
    return ToolResult(content=[TextContent(text=json.dumps(data, indent=2))])

def _get_wb(path):
    if not os.path.exists(path):
        raise FileNotFoundError(f"File not found: {path}")
    return load_workbook(path)

async def render_template(arguments: dict) -> ToolResult:
    """
    Replace {{key}} placeholders in the workbook.
    Args:
        replacements: dict {"name": "John", "date": "2024-01-01"}
    """
    try:
        path = arguments['file_path']
        replacements = arguments['replacements']
        sheet_name = arguments.get('sheet_name')
        
        wb = _get_wb(path)
        ws = wb[sheet_name] if sheet_name else wb.active
        
        count = 0
        
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and "{{" in cell.value:
                    val = cell.value
                    changed = False
                    for k, v in replacements.items():
                        placeholder = "{{" + k + "}}"
                        if placeholder in val:
                            val = val.replace(placeholder, str(v))
                            changed = True
                    
                    if changed:
                        cell.value = val
                        count += 1
                        
        wb.save(path)
        wb.close()
        return dict_to_result({"status": "template_rendered", "replacements_count": count})
    except Exception as e:
        return ToolResult(isError=True, content=[TextContent(text=str(e))])

async def add_sparkline(arguments: dict) -> ToolResult:
    return ToolResult(content=[TextContent(text="Sparkline support is limited in current OpenPyXL version.")])
