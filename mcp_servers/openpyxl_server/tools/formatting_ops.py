
from openpyxl import load_workbook
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule, Rule
from openpyxl.styles import PatternFill, Font, Border, Side
from openpyxl.styles.differential import DifferentialStyle
from shared.mcp.protocol import ToolResult, TextContent
from shared.logging.main import get_logger
import os
import json

logger = get_logger(__name__)

def dict_to_result(data: dict) -> ToolResult:
    return ToolResult(content=[TextContent(text=json.dumps(data, indent=2))])

def _get_wb(path):
    if not os.path.exists(path):
        raise FileNotFoundError(f"File not found: {path}")
    return load_workbook(path)

# ============================================================================
# Conditional Formatting
# ============================================================================

async def add_color_scale(arguments: dict) -> ToolResult:
    """Add 2 or 3 color scale rule."""
    try:
        path = arguments['file_path']
        range_string = arguments['range_string']
        start_color = arguments.get('start_color', 'F8696B') # Red
        mid_color = arguments.get('mid_color', 'FFEB84')   # Yellow
        end_color = arguments.get('end_color', '63BE7B')   # Green
        sheet_name = arguments.get('sheet_name')
        
        wb = _get_wb(path)
        ws = wb[sheet_name] if sheet_name else wb.active
        
        # 3-Color Scale
        rule = ColorScaleRule(start_type='min', start_color=start_color,
                              mid_type='percentile', mid_value=50, mid_color=mid_color,
                              end_type='max', end_color=end_color)
                              
        ws.conditional_formatting.add(range_string, rule)
        
        wb.save(path)
        wb.close()
        return dict_to_result({"status": "color_scale_added", "range": range_string})
    except Exception as e:
        return ToolResult(isError=True, content=[TextContent(text=str(e))])

async def add_data_bar(arguments: dict) -> ToolResult:
    """Add Data Bar rule."""
    try:
        path = arguments['file_path']
        range_string = arguments['range_string']
        color = arguments.get('color', '638EC6')
        sheet_name = arguments.get('sheet_name')
        
        wb = _get_wb(path)
        ws = wb[sheet_name] if sheet_name else wb.active
        
        from openpyxl.formatting.rule import DataBarRule
        rule = DataBarRule(start_type='min', end_type='max', color=color)
        ws.conditional_formatting.add(range_string, rule)
        
        wb.save(path)
        wb.close()
        return dict_to_result({"status": "data_bar_added", "range": range_string})
    except Exception as e:
        return ToolResult(isError=True, content=[TextContent(text=str(e))])

async def add_highlight_rule(arguments: dict) -> ToolResult:
    """
    Add a Cell Is rule (e.g. Greater Than 100 -> Red Fill).
    operator: 'greaterThan', 'lessThan', 'equal', 'between'
    formula: [value] or [val1, val2]
    """
    try:
        path = arguments['file_path']
        range_string = arguments['range_string']
        operator = arguments['operator']
        formula = arguments['formula'] # List of strings/ints
        sheet_name = arguments.get('sheet_name')
        
        # Style arguments
        fill_color = arguments.get('fill_color')
        font_color = arguments.get('font_color')
        bold = arguments.get('bold', False)
        
        wb = _get_wb(path)
        ws = wb[sheet_name] if sheet_name else wb.active
        
        # Build Style
        dxf = DifferentialStyle()
        if font_color or bold:
             dxf.font = Font(color=font_color, bold=bold)
        if fill_color:
             dxf.fill = PatternFill(bgColor=fill_color)
             
        rule = CellIsRule(operator=operator, formula=formula, stopIfTrue=True, dxf=dxf)
        ws.conditional_formatting.add(range_string, rule)
        
        wb.save(path)
        wb.close()
        return dict_to_result({"status": "highlight_rule_added", "range": range_string, "operator": operator})
    except Exception as e:
        return ToolResult(isError=True, content=[TextContent(text=str(e))])

# ============================================================================
# Page Setup
# ============================================================================

async def set_page_setup(arguments: dict) -> ToolResult:
    """Set print area, orientation, paper size."""
    try:
        path = arguments['file_path']
        orientation = arguments.get('orientation') # 'portrait', 'landscape'
        paper_size = arguments.get('paper_size') # index
        fit_to_width = arguments.get('fit_to_width') # int (pages)
        fit_to_height = arguments.get('fit_to_height')
        sheet_name = arguments.get('sheet_name')
        
        wb = _get_wb(path)
        ws = wb[sheet_name] if sheet_name else wb.active
        
        if orientation: ws.page_setup.orientation = orientation
        if paper_size: ws.page_setup.paperSize = paper_size
        if fit_to_width is not None: ws.page_setup.fitToWidth = fit_to_width
        if fit_to_height is not None: ws.page_setup.fitToHeight = fit_to_height
        
        wb.save(path)
        wb.close()
        return dict_to_result({"status": "page_setup_updated"})
    except Exception as e:
        return ToolResult(isError=True, content=[TextContent(text=str(e))])

async def set_print_area(arguments: dict) -> ToolResult:
    try:
        path = arguments['file_path']
        print_area = arguments['print_area'] # 'A1:M30'
        sheet_name = arguments.get('sheet_name')
        
        wb = _get_wb(path)
        ws = wb[sheet_name] if sheet_name else wb.active
        
        ws.print_area = print_area
        
        wb.save(path)
        wb.close()
        return dict_to_result({"status": "print_area_set", "area": print_area})
    except Exception as e:
        return ToolResult(isError=True, content=[TextContent(text=str(e))])
