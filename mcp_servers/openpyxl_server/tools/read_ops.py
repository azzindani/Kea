
from openpyxl import load_workbook
from shared.mcp.protocol import ToolResult, TextContent
from shared.logging.main import get_logger
import os
import json
import re

logger = get_logger(__name__)

def dict_to_result(data: dict) -> ToolResult:
    return ToolResult(content=[TextContent(text=json.dumps(data, indent=2, default=str))])

def _get_wb(path, read_only=True, data_only=True):
    if not os.path.exists(path):
        raise FileNotFoundError(f"File not found: {path}")
    return load_workbook(path, read_only=read_only, data_only=data_only)

# ============================================================================
# Bulk Reading Strategy
# ============================================================================

async def read_excel_multitalent(arguments: dict) -> ToolResult:
    """
    Super-Tool: Flexible reader.
    Args:
        file_path: str
        sheet_name: str (optional, default active)
        range_string: str (optional, e.g. 'A1:C10')
        read_mode: 'values' | 'formulas' (default 'values')
        header_row: int (optional, if set, returns list of dicts)
    """
    try:
        path = arguments['file_path']
        sheet_name = arguments.get('sheet_name')
        range_string = arguments.get('range_string')
        read_mode = arguments.get('read_mode', 'values')
        header_row_idx = arguments.get('header_row') # 1-based index
        
        data_only = (read_mode == 'values')
        wb = _get_wb(path, read_only=True, data_only=data_only)
        
        ws = wb[sheet_name] if sheet_name else wb.active
        
        # Determine iterator
        if range_string:
            rows_iter = ws[range_string]
            # In read_only, ws['A1:B2'] might not work as iterator directly for values_only?
            # Actually ws[range] returns tuple of tuples of cells.
            # We map it manually safely.
            data = []
            for row in rows_iter:
                # Row is a tuple of cells
                row_vals = [cell.value for cell in row]
                data.append(row_vals)
        else:
            # Full sheet (iter_rows is efficient)
            data = list(ws.iter_rows(values_only=True))
            
        wb.close()
        
        # Process Headers
        if header_row_idx and len(data) >= header_row_idx:
            # 1-based index, so index-1
            headers = data[header_row_idx - 1]
            body = data[header_row_idx:]
            
            structured = []
            for row in body:
                item = {}
                for h, v in zip(headers, row):
                    if h: item[str(h)] = v
                structured.append(item)
            return dict_to_result({"data": structured, "count": len(structured)})
            
        return dict_to_result({"data": data, "count": len(data)})

    except Exception as e:
        return ToolResult(isError=True, content=[TextContent(text=str(e))])

async def read_cell_value(arguments: dict) -> ToolResult:
    try:
        path = arguments['file_path']
        sheet = arguments.get('sheet_name')
        coord = arguments['cell'] # 'A1'
        
        wb = _get_wb(path, read_only=True)
        ws = wb[sheet] if sheet else wb.active
        val = ws[coord].value
        wb.close()
        return dict_to_result({"value": val, "cell": coord})
    except Exception as e:
        return ToolResult(isError=True, content=[TextContent(text=str(e))])

async def search_excel_content(arguments: dict) -> ToolResult:
    """
    Search for a value or regex across all sheets.
    """
    try:
        path = arguments['file_path']
        query = arguments['query']
        is_regex = arguments.get('is_regex', False)
        
        wb = _get_wb(path, read_only=True, data_only=True)
        
        results = []
        pattern = re.compile(query, re.IGNORECASE) if is_regex else None
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    val = str(cell.value) if cell.value is not None else ""
                    match = False
                    if is_regex:
                        if pattern.search(val): match = True
                    else:
                        if query.lower() in val.lower(): match = True
                        
                    if match:
                        results.append({
                            "sheet": sheet_name,
                            "cell": cell.coordinate,
                            "value": cell.value
                        })
                        if len(results) > 100: # Safety limit
                             break
                if len(results) > 100: break
                
        wb.close()
        return dict_to_result({"matches": results, "count": len(results)})
    except Exception as e:
        return ToolResult(isError=True, content=[TextContent(text=str(e))])

# Helpers for Rows/Cols
async def read_row_values(arguments: dict) -> ToolResult:
    # ... straightforward implementation ...
    # Skipping detailed impl for brevity of this turn, relying on multitalent
    # user asked for "list what tools can be created". I will verify the plan covers them.
    # I'll implement a catch-all reader for rows/cols in multitalent or specific tools if requested.
    # Let's add read_range explicitly as it's cleaner.
    pass

async def read_range_values(arguments: dict) -> ToolResult:
    """Wrapper for specific range read."""
    # Reuse multitalent logic
    arguments['read_mode'] = 'values'
    return await read_excel_multitalent(arguments)
