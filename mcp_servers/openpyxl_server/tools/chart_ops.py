
from openpyxl import load_workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from shared.mcp.protocol import ToolResult, TextContent
from shared.logging.main import get_logger
import os
import json

logger = get_logger(__name__)

def dict_to_result(data: dict) -> ToolResult:
    return ToolResult(content=[TextContent(text=json.dumps(data, indent=2))])

def _get_wb(path):
    if not os.path.exists(path):
        raise FileNotFoundError(f"File not found: {path}")
    return load_workbook(path)

async def create_chart(arguments: dict) -> ToolResult:
    """
    Create a chart from data range.
    Args:
        type: 'bar', 'line', 'pie', 'col'
        data_range: 'Sheet1!A1:C10' OR just range 'A1:C10' (uses active sheet if no prefix)
        title: str
        anchor: 'E5' (Top-left cell for chart)
    """
    try:
        path = arguments['file_path']
        chart_type = arguments['type'].lower()
        data_range = arguments['data_range']
        title = arguments.get('title', 'Chart')
        anchor = arguments.get('anchor', 'E1')
        sheet_name = arguments.get('sheet_name')
        
        wb = _get_wb(path)
        ws = wb[sheet_name] if sheet_name else wb.active
        
        # Parse range
        # openpyxl Reference expects min_col, min_row, etc.
        # But we can try to use ws[range] to get boundaries?
        # Or parse string "A1:C10"
        
        if '!' in data_range:
            # Sheet included in range string, potentially different from target sheet?
            # For simplicity, assume data is on same sheet or handle parsing
            pass

        # Helper to parse range string to min/max row/col
        # This is tricky without a helper. Openpyxl has utils.
        from openpyxl.utils.cell import range_boundaries
        min_col, min_row, max_col, max_row = range_boundaries(data_range)
        
        data = Reference(ws, min_col=min_col, min_row=min_row, max_col=max_col, max_row=max_row)
        
        c = None
        if 'bar' in chart_type or 'col' in chart_type:
            c = BarChart()
            if 'col' in chart_type: c.type = "col"
            else: c.type = "bar"
        elif 'line' in chart_type:
            c = LineChart()
        elif 'pie' in chart_type:
            c = PieChart()
        else:
            wb.close()
            return ToolResult(isError=True, content=[TextContent(text=f"Unknown chart type: {chart_type}")])
            
        c.add_data(data, titles_from_data=True) # Assume row 1 is titles? Make optional?
        c.title = title
        
        ws.add_chart(c, anchor)
        
        wb.save(path)
        wb.close()
        return dict_to_result({"status": "chart_added", "type": chart_type, "anchor": anchor})
    except Exception as e:
        return ToolResult(isError=True, content=[TextContent(text=str(e))])

async def delete_chart(arguments: dict) -> ToolResult:
    # OpenPyXL doesn't easily support deleting charts by name unless we iterate _charts
    # This is advanced.
    return ToolResult(content=[TextContent(text="feature_not_implemented_yet")])
