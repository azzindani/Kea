
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.comments import Comment
from shared.mcp.protocol import ToolResult, TextContent
from shared.logging.main import get_logger
import os
import json

logger = get_logger(__name__)

def dict_to_result(data: dict) -> ToolResult:
    return ToolResult(content=[TextContent(text=json.dumps(data, indent=2))])

def _get_wb(path):
    if not os.path.exists(path):
        raise FileNotFoundError(f"File not found: {path}")
    return load_workbook(path)

async def add_image_to_sheet(arguments: dict) -> ToolResult:
    try:
        path = arguments['file_path']
        image_path = arguments['image_path']
        anchor = arguments.get('anchor', 'A1')
        sheet_name = arguments.get('sheet_name')
        
        if not os.path.exists(image_path):
             return ToolResult(isError=True, content=[TextContent(text=f"Image not found: {image_path}")])

        wb = _get_wb(path)
        ws = wb[sheet_name] if sheet_name else wb.active
        
        img = Image(image_path)
        ws.add_image(img, anchor)
        
        wb.save(path)
        wb.close()
        return dict_to_result({"status": "image_added", "image": image_path, "anchor": anchor})
    except Exception as e:
        return ToolResult(isError=True, content=[TextContent(text=str(e))])

async def add_comment(arguments: dict) -> ToolResult:
    try:
        path = arguments['file_path']
        cell = arguments['cell']
        text = arguments['text']
        author = arguments.get('author', 'Project Agent')
        sheet_name = arguments.get('sheet_name')
        
        wb = _get_wb(path)
        ws = wb[sheet_name] if sheet_name else wb.active
        
        comment = Comment(text, author)
        ws[cell].comment = comment
        
        wb.save(path)
        wb.close()
        return dict_to_result({"status": "comment_added", "cell": cell})
    except Exception as e:
        return ToolResult(isError=True, content=[TextContent(text=str(e))])

async def read_comments(arguments: dict) -> ToolResult:
    try:
        path = arguments['file_path']
        sheet_name = arguments.get('sheet_name')
        
        wb = _get_wb(path)
        ws = wb[sheet_name] if sheet_name else wb.active
        
        comments = []
        # ws.comments is not directly iterable in all versions??
        # Iterate all cells? Slow.
        # But wait, openpyxl doesn't expose a list of comments easily without iterating cells.
        # Optimized: only iter cells with comments if we track them? No.
        # Fallback: simple iteration for now, restrict to used range.
        
        for row in ws.iter_rows():
            for cell in row:
                if cell.comment:
                    comments.append({
                        "cell": cell.coordinate,
                        "text": cell.comment.text,
                        "author": cell.comment.author
                    })
        
        wb.close()
        return dict_to_result({"comments": comments})
    except Exception as e:
        return ToolResult(isError=True, content=[TextContent(text=str(e))])
